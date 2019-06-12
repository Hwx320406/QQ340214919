from openpyxl import load_workbook
# from tool.parametric_switch import edu_data

class Excel_data(object):

    def fetch_excel_testData(self,path,subpage,testname):
        #获取所有exlce子页
        pageList = load_workbook(path)
        #获取当前测试子页
        page =pageList[subpage]
        # A 为子页的A行
        for i in range(len(page['A'])):
            testline = page['A'][i].value
            # 判断exlce表A行里的测试编码与脚本测试方法对应的数据
            if testline == testname:
                #获取row 行column列的值
                testData=page.cell(row=page['A'][i].row,column=4).value
                if testData is not None:
                    # 获取一个字符串的索引
                    stuff = testData.find(',')
                    if stuff != -1:
                        data=testData.split(',')

                        return data


    def edu_data(self,path, subpage, testname, clom,*args,**kwargs):
        # 获取excel测试数据
        data1 = self.fetch_excel_testData(path, subpage, testname)
        #登录数据
        if clom == 1:
            data = {'username': data1[0], 'password': data1[1]}
            return data
        #添加教师数据
        elif clom ==2:
            data ={'username':data1[0],'realname':data1[1],'password':data1[2],'sex':data1[3],'roleid':data1[4],
            'orid1':data1[5],'email':data1[6],'phone':data1[7],'location_p':data1[8],'location_c':data1[9],
            'location_a':data1[10], 'address':data1[11],'introduce':data1[12],'type':data1[13]}

            return data